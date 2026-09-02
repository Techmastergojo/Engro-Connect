import openpyxl
import json
import os
from datetime import datetime
from collections import defaultdict

def safe_float(v, default=0.0):
    try:
        if v is None: return default
        return float(v)
    except:
        return default

def clean_val(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    return str(v).split()[0] if ' ' in str(v) else str(v)

def parse_all():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_perf = os.path.join(base_dir, "C4 Overall Performance Aug-2026 (3).xlsx")
    if not os.path.exists(file_perf):
        file_perf = r"c:\Users\hamza\OneDrive\Desktop\Engro Connect\C4 Overall Performance Aug-2026 (3).xlsx"

    file_fuel = os.path.join(base_dir, "C-4 Daily Deodar Fuel Activity Report 30th August-2026.xlsx")
    if not os.path.exists(file_fuel):
        file_fuel = r"c:\Users\hamza\OneDrive\Desktop\Engro Connect\C-4 Daily Deodar Fuel Activity Report 30th August-2026.xlsx"

    print(f"Loading Performance workbook from {file_perf}...")
    wb_perf = openpyxl.load_workbook(file_perf, read_only=True, data_only=True)

    # 1. Parse E2E & Deodar NAR (all 31 active days in August)
    ws = wb_perf['E2E & Deodar NAR']
    rows = list(ws.iter_rows(values_only=True))
    dates = []
    for cell in rows[0][1:]:
        if cell is not None:
            dates.append(clean_val(cell))

    deodar_daily = {}
    e2e_daily = {}
    for idx, d in enumerate(dates):
        if idx + 1 < len(rows[1]) and rows[1][idx + 1] is not None:
            try: deodar_daily[d] = round(float(rows[1][idx + 1]) * 100, 2)
            except: pass
        if idx + 1 < len(rows[2]) and rows[2][idx + 1] is not None:
            try: e2e_daily[d] = round(float(rows[2][idx + 1]) * 100, 2)
            except: pass

    active_dates = list(deodar_daily.keys())
    avg_deodar = round(sum(deodar_daily.values()) / max(len(deodar_daily), 1), 2)
    avg_e2e = round(sum(e2e_daily.values()) / max(len(e2e_daily), 1), 2)
    print(f"Found {len(active_dates)} active dates in August.")

    # 2. Parse DateWiseDT (MBU daily NAR)
    ws = wb_perf['DateWiseDT']
    rows = list(ws.iter_rows(values_only=True))
    mbu_list = [str(r).strip() for r in rows[0][1:9] if r is not None]
    mbu_daily_nar = {mbu: {} for mbu in mbu_list}
    for r in rows[1:]:
        if not r[0]: continue
        d = clean_val(r[0])
        if d not in active_dates: continue
        for idx, mbu in enumerate(mbu_list):
            if idx + 1 < len(r) and r[idx + 1] is not None:
                try: mbu_daily_nar[mbu][d] = round(float(r[idx + 1]) * 100, 2)
                except: pass

    # 3. Parse MBUWiseDT (MBU daily Downtime in hours)
    ws = wb_perf['MBUWiseDT']
    rows = list(ws.iter_rows(values_only=True))
    mbu_daily_dt = {mbu: {} for mbu in mbu_list}
    for r in rows[1:9]:
        mbu = str(r[0]).strip()
        if mbu in mbu_daily_dt:
            for idx, d in enumerate(active_dates):
                if idx + 1 < len(r) and r[idx + 1] is not None:
                    mbu_daily_dt[mbu][d] = round(safe_float(r[idx + 1]) / 60, 1)

    # 4. Parse MBUWiseContribution (Official TDT & TNAR)
    ws = wb_perf['MBUWiseContribution']
    mbu_totals = {}
    for r in list(ws.iter_rows(values_only=True))[1:15]:
        if r[0] and str(r[0]).startswith('C4-'):
            mbu = str(r[0]).strip()
            tdt_min = round(safe_float(r[1]) / 60, 1)
            tnar = round(safe_float(r[2]) * 100, 2)
            mbu_totals[mbu] = {'tdtMinutes': tdt_min, 'tnar': tnar}

    # 5. Elite-Platinum
    ws = wb_perf['Elite-Platinum Graphs']
    rows = list(ws.iter_rows(values_only=True))
    elite_platinum = {
        'elite': {'nar': round(float(rows[1][1]) * 100, 2) if rows[1][1] else 0, 'totalSites': int(rows[1][2]) if rows[1][2] else 0},
        'platinum': {'nar': round(float(rows[1][4]) * 100, 2) if rows[1][4] else 0, 'totalSites': int(rows[1][5]) if rows[1][5] else 0}
    }

    # 6. Parse 4G Month Wise History (Jan-Jul 2026 site history)
    print("Parsing 4G Month Wise History...")
    ws = wb_perf['4G Month Wise History']
    history_map = {}
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul']
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not r[0]: continue
        site_code = str(r[0]).strip().upper()
        months = {}
        vals = []
        for idx, m_name in enumerate(month_names):
            col_idx = 11 + idx
            if col_idx < len(r) and r[col_idx] is not None:
                try:
                    v = round(float(r[col_idx]), 2)
                    months[m_name] = v
                    vals.append(v)
                except: pass
        avg6m = round(sum(vals[-6:]) / len(vals[-6:]), 2) if vals else 100.0
        history_map[site_code] = {'months': months, 'avg6m': avg6m}
    print(f"Parsed 4G history for {len(history_map)} sites.")

    # 7. Parse Consolidated RSL Aug-26 (Outage reasons per site and period)
    print("Parsing Consolidated RSL Aug-26 outages...")
    ws = wb_perf['Consolidated RSL Aug-26']
    site_outages = defaultdict(lambda: {
        '3d': defaultdict(float),
        '7d': defaultdict(float),
        '15d': defaultdict(float),
        '30d': defaultdict(float),
        'domains': defaultdict(float),
        'totalDt': 0.0,
        'count': 0
    })

    # Overall C4 and MBU outage aggregations
    c4_outage_reasons = {'3d': defaultdict(float), '7d': defaultdict(float), '15d': defaultdict(float), '30d': defaultdict(float)}
    mbu_outage_reasons = defaultdict(lambda: {'3d': defaultdict(float), '7d': defaultdict(float), '15d': defaultdict(float), '30d': defaultdict(float)})

    d3_start = '2026-08-29'
    d7_start = '2026-08-25'
    d15_start = '2026-08-17'

    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not r[21]: continue
        site_code = str(r[21]).strip().upper()
        mbu = str(r[17]).strip() if len(r) > 17 and r[17] else ''
        dt = float(r[7]) if r[7] is not None else 0.0
        reason = str(r[29]).strip() if r[29] else 'Other'
        domain = str(r[33]).strip() if len(r) > 33 and r[33] else 'Other'
        date_str = str(r[5])[:10] if r[5] else ''

        so = site_outages[site_code]
        so['30d'][reason] += dt
        so['domains'][domain] += dt
        so['totalDt'] += dt
        so['count'] += 1

        c4_outage_reasons['30d'][reason] += dt
        if mbu: mbu_outage_reasons[mbu]['30d'][reason] += dt

        if date_str >= d15_start:
            so['15d'][reason] += dt
            c4_outage_reasons['15d'][reason] += dt
            if mbu: mbu_outage_reasons[mbu]['15d'][reason] += dt
        if date_str >= d7_start:
            so['7d'][reason] += dt
            c4_outage_reasons['7d'][reason] += dt
            if mbu: mbu_outage_reasons[mbu]['7d'][reason] += dt
        if date_str >= d3_start:
            so['3d'][reason] += dt
            c4_outage_reasons['3d'][reason] += dt
            if mbu: mbu_outage_reasons[mbu]['3d'][reason] += dt

    print(f"Aggregated outages for {len(site_outages)} sites.")

    # 8. Parse Site NAR-Day (all August days + Total NAR)
    print("Parsing Site NAR-Day...")
    ws = wb_perf['Site NAR-Day']
    site_nar_rows = list(ws.iter_rows(values_only=True))
    site_nar_dates = [clean_val(c) for c in site_nar_rows[0][8:39] if c is not None]
    sites_nar = []

    for r in site_nar_rows[1:]:
        if not r[1]: continue
        site_code = str(r[1]).strip()
        site_code_upper = site_code.upper()
        site_name = str(r[2]).strip() if r[2] else ''
        mbu = str(r[4]).strip() if r[4] else ''

        # Daily August NAR
        daily = {}
        for idx, d in enumerate(site_nar_dates):
            if 8 + idx < len(r) and r[8 + idx] is not None:
                daily[d] = round(safe_float(r[8 + idx]) * 100, 1)

        daily_vals = list(daily.values())

        # Total NAR from Col 39 (Average NAR)
        total_nar = 100.0
        if len(r) > 39 and r[39] is not None:
            total_nar = round(safe_float(r[39]) * 100, 2)
        elif daily_vals:
            total_nar = round(sum(daily_vals) / len(daily_vals), 2)

        # 3d, 7d, 15d, 30d NAR
        nar3d = round(sum(daily_vals[-3:]) / max(len(daily_vals[-3:]), 1), 2) if daily_vals else total_nar
        nar7d = round(sum(daily_vals[-7:]) / max(len(daily_vals[-7:]), 1), 2) if daily_vals else total_nar
        nar15d = round(sum(daily_vals[-15:]) / max(len(daily_vals[-15:]), 1), 2) if daily_vals else total_nar
        nar30d = total_nar

        # 6-Month history and trend
        hist_entry = history_map.get(site_code_upper, None)
        hist_months = dict(hist_entry['months']) if hist_entry else {}
        hist_months['Aug'] = total_nar
        
        # 6m average combining historical + August
        h_vals = list(hist_months.values())[-6:]
        nar6m = round(sum(h_vals) / len(h_vals), 2) if h_vals else total_nar

        # Outage stats for this site
        outage_entry = site_outages.get(site_code_upper, None)
        outage_stats = None
        if outage_entry:
            def top_reasons_dict(d, limit=6):
                sorted_r = sorted(d.items(), key=lambda x: -x[1])[:limit]
                return {k: round(v, 1) for k, v in sorted_r}

            outage_stats = {
                'totalDt': round(outage_entry['totalDt'], 1),
                'count': outage_entry['count'],
                'reasons3d': top_reasons_dict(outage_entry['3d']),
                'reasons7d': top_reasons_dict(outage_entry['7d']),
                'reasons15d': top_reasons_dict(outage_entry['15d']),
                'reasons30d': top_reasons_dict(outage_entry['30d']),
                'domains': {k: round(v, 1) for k, v in sorted(outage_entry['domains'].items(), key=lambda x: -x[1])[:4]}
            }

        sites_nar.append({
            'code': site_code,
            'name': site_name,
            'mbu': mbu,
            'avgNar': total_nar,
            'totalNar': total_nar,
            'nar3d': nar3d,
            'nar7d': nar7d,
            'nar15d': nar15d,
            'nar30d': nar30d,
            'nar6m': nar6m,
            'history6m': hist_months,
            'outageStats': outage_stats,
            'daily': daily
        })

    # 9. Outage Categories master list
    ws = wb_perf['Outage Category']
    outage_categories = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if r[0] and r[1]:
            outage_categories.append({'reason': str(r[0]).strip(), 'domain': str(r[1]).strip()})

    wb_perf.close()
    print(f"Parsed {len(sites_nar)} sites with complete multi-period NAR and outage data.")

    # 10. Parse Fuel Data
    print(f"Loading Fuel workbook from {file_fuel}...")
    wb_fuel = openpyxl.load_workbook(file_fuel, read_only=True, data_only=True)

    # 10a. Fuel Summary
    ws = wb_fuel['Summary']
    fuel_summary_daily = []
    for r in list(ws.iter_rows(values_only=True))[3:]:
        if len(r) > 7 and r[1] is not None:
            d = clean_val(r[1])
            if not isinstance(d, str) or not d.startswith('2026-'): continue
            fuel_summary_daily.append({
                'date': d,
                'scratching': safe_float(r[2], 0),
                'pouring': safe_float(r[5], 0),
                'inhand': safe_float(r[7], 0)
            })

    # 10b. MBU Wise Fuel Detail
    ws = wb_fuel['Fuel Detail MBU Wise ']
    fuel_mbu_rows = list(ws.iter_rows(values_only=True))
    fuel_mbu_headers = [str(x).strip() for x in fuel_mbu_rows[1][2:10] if x is not None]
    fuel_mbu_daily = []
    for r in fuel_mbu_rows[2:]:
        if r[1] is not None:
            d = clean_val(r[1])
            entry = {'date': d}
            total = 0
            for idx, mbu in enumerate(fuel_mbu_headers):
                if 2 + idx < len(r) and r[2 + idx] is not None:
                    val = safe_float(r[2 + idx], 0)
                    entry[mbu] = val
                    total += val
            entry['total'] = total
            fuel_mbu_daily.append(entry)

    # 10c. Vehicle Wise Fuel Detail
    ws = wb_fuel['Vehicle Wise Fuel Detail']
    vehicle_rows = list(ws.iter_rows(values_only=True))
    v_dates = [clean_val(c) for c in vehicle_rows[2][5:] if c is not None]
    vehicles = []
    for r in vehicle_rows[3:]:
        if not r[1]: continue
        v_daily = {}
        v_total = 0
        for idx, d in enumerate(v_dates):
            if 5 + idx < len(r) and r[5 + idx] is not None:
                val = safe_float(r[5 + idx], 0)
                v_daily[d] = val
                v_total += val
        vehicles.append({
            'vehicleNo': str(r[1]).strip(),
            'fsoName': str(r[2]).strip() if r[2] else '',
            'area': str(r[3]).strip() if r[3] else '',
            'mbu': str(r[4]).strip() if r[4] else '',
            'totalFuel': v_total,
            'daily': v_daily
        })

    # 10d. Site DG Profiles
    ws = wb_fuel['Site Detail']
    site_dg_profiles = []
    site_dg_map = {}
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[0]: continue
        code = str(r[0]).strip()
        cf = None
        if len(r) > 9 and r[9] is not None:
            try: cf = float(r[9])
            except: cf = None
        profile = {
            'siteCode': code,
            'siteName': str(r[1]).strip() if len(r) > 1 and r[1] else '',
            'mbu': str(r[2]).strip() if len(r) > 2 and r[2] else '',
            'flmVendor': str(r[3]).strip() if len(r) > 3 and r[3] else '',
            'secVendor': str(r[4]).strip() if len(r) > 4 and r[4] else '',
            'dgKva': str(r[6]).strip() if len(r) > 6 and r[6] is not None else '',
            'engineBrand': str(r[7]).strip() if len(r) > 7 and r[7] else '',
            'dgBrand': str(r[8]).strip() if len(r) > 8 and r[8] else '',
            'consumptionFactor': cf
        }
        site_dg_profiles.append(profile)
        site_dg_map[code] = profile

    # 10e. Site Pouring Logs
    ws = wb_fuel['\u26fdPouring']
    pouring_logs = []
    site_fuel_summary = {}
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if not r[1]: continue
        code = str(r[1]).strip()
        p_date = clean_val(r[8])
        poured = safe_float(r[10])
        bal = safe_float(r[9])
        entry = {
            'siteCode': code,
            'siteName': str(r[2]).strip() if r[2] else '',
            'mbu': str(r[4]).strip() if r[4] else '',
            'vendor': str(r[5]).strip() if r[5] else '',
            'siteStatus': str(r[6]).strip() if r[6] else '',
            'visitNature': str(r[7]).strip() if r[7] else '',
            'date': p_date,
            'fuelBal': bal,
            'fuelPoured': poured,
            'vehicleNo': str(r[19]).strip() if len(r) > 19 and r[19] else '',
            'createdBy': str(r[16]).strip() if len(r) > 16 and r[16] else ''
        }
        pouring_logs.append(entry)
        
        if code not in site_fuel_summary:
            site_fuel_summary[code] = {
                'siteCode': code,
                'siteName': entry['siteName'],
                'mbu': entry['mbu'],
                'totalPoured': 0,
                'visitsCount': 0,
                'lastVisitDate': p_date,
                'lastFuelBal': bal,
                'dgProfile': site_dg_map.get(code, None)
            }
        site_fuel_summary[code]['totalPoured'] += poured
        site_fuel_summary[code]['visitsCount'] += 1

    # Include any DG sites that didn't have pouring activity in site_fuel_summary
    for code, dg in site_dg_map.items():
        if code not in site_fuel_summary:
            site_fuel_summary[code] = {
                'siteCode': code,
                'siteName': dg['siteName'],
                'mbu': dg['mbu'],
                'totalPoured': 0,
                'visitsCount': 0,
                'lastVisitDate': 'N/A',
                'lastFuelBal': 0,
                'dgProfile': dg
            }

    # 10f. Scratching Cards Logs
    ws = wb_fuel['\U0001f4b3Scratching']
    scratching_logs = []
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if not r[1]: continue
        scratching_logs.append({
            'cardCode': str(r[1]).strip(),
            'fuelCompany': str(r[2]).strip() if len(r) > 2 and r[2] else '',
            'vendor': str(r[3]).strip() if len(r) > 3 and r[3] else '',
            'cardName': str(r[4]).strip() if len(r) > 4 and r[4] else '',
            'cardNumber': str(r[5]).strip() if len(r) > 5 and r[5] else '',
            'cardLimit': safe_float(r[6], 0) if len(r) > 6 else 0,
            'mbu': str(r[7]).strip() if len(r) > 7 and r[7] else '',
            'region': str(r[8]).strip() if len(r) > 8 and r[8] else '',
            'date': clean_val(r[9]),
            'quantity': safe_float(r[11], 0) if len(r) > 11 else 0,
            'station': str(r[12]).strip() if len(r) > 12 and r[12] else ''
        })

    wb_fuel.close()

    def format_reasons(d, limit=8):
        return {k: round(v, 1) for k, v in sorted(d.items(), key=lambda x: -x[1])[:limit]}

    c4_reasons_formatted = {p: format_reasons(c4_outage_reasons[p]) for p in ['3d', '7d', '15d', '30d']}
    mbu_reasons_formatted = {
        mbu: {p: format_reasons(mbu_outage_reasons[mbu][p]) for p in ['3d', '7d', '15d', '30d']}
        for mbu in mbu_list
    }

    analytics_data = {
        'dates': active_dates,
        'mbuList': mbu_list,
        'nar': {
            'wholeC4': {
                'deodarDaily': deodar_daily,
                'e2eDaily': e2e_daily,
                'avgDeodar': avg_deodar,
                'avgE2E': avg_e2e,
            },
            'mbuDaily': mbu_daily_nar,
            'mbuDailyDt': mbu_daily_dt,
            'mbuTotals': mbu_totals,
            'elitePlatinum': elite_platinum,
            'outageCategories': outage_categories,
            'c4OutageReasons': c4_reasons_formatted,
            'mbuOutageReasons': mbu_reasons_formatted,
            'sites': sites_nar
        },
        'fuel': {
            'summaryDaily': fuel_summary_daily,
            'totalScratched': sum(x['scratching'] for x in fuel_summary_daily),
            'totalPoured': sum(x['pouring'] for x in fuel_summary_daily),
            'mbuDaily': fuel_mbu_daily,
            'vehicles': vehicles,
            'pouringLogs': pouring_logs,
            'scratchingLogs': scratching_logs,
            'siteDgProfiles': site_dg_profiles,
            'siteFuelList': list(site_fuel_summary.values())
        }
    }

    out_file = os.path.join(base_dir, "src", "analyticsData.ts")
    with open(out_file, "w", encoding="utf-8") as f:
        f.write("// @ts-nocheck\nexport const analyticsData = ")
        json.dump(analytics_data, f, separators=(',', ':'))
        f.write(";\n")

    print(f"Analytics Data exported successfully to {out_file} (Size: {os.path.getsize(out_file) / 1024:.1f} KB)")

if __name__ == '__main__':
    parse_all()
